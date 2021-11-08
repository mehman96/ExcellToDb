
from pathlib import Path  # Standard Python Module
from openpyxl import load_workbook, Workbook  # pip install openpyxl


# -- 1.STEP
# Get all excel file paths from given directory
SOURCE_DIR = "Daily_Reports" # e.g. r"C:\Users\Username\Desktop\Sample Files"
excel_files = list(Path(SOURCE_DIR).glob("*.xlsx"))

# -- 2.STEP:
# Iterate over all Excel files from step 1,
# access the worksheet and store the values in a dictionary
# values_excel_files = {['2021-01-01'] : [1,2,3, ..],
#                       ['2021-01-02'] : [1,2,3, ..]}
values_excel_files = {}
for excel_file in excel_files:
    report_date = excel_file.stem.replace("_Report", "")
    wb = load_workbook(filename=excel_file, read_only=True)
    rng = wb["Sheet1"]["B2":"B19"]
    rng_values = []
    for cells in rng:
        for cell in cells:
            rng_values.append(cell.value)
    values_excel_files[report_date] = rng_values

# -- 3.STEP:
# a) Iterate over all worksheets in the master workbook
# b) For each worksheet, iterate over defined Excel range (dates)
# c) If date matches with the key of dictionary (values_excel_files) then insert values & save workbook
wb = load_workbook(filename="Masterfile_Template.xlsx")
for ws in wb.worksheets:
    clm = "B"
    first_row = 3
    last_row = len(ws[clm])
    rng = ws[f"{clm}{first_row}:{clm}{last_row}"]
    for cells in rng:
        for cell in cells:
            if cell.value in values_excel_files:
                # Iterate over values (list inside the dictionary) and write values to column
                for i, value in enumerate(values_excel_files[cell.value]):
                    cell.offset(row=0, column=i + 1).value = value
wb.save("data.xlsx")